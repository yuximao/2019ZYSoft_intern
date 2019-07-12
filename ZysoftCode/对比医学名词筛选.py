# 将每篇文章与医学字典比对，看其中含有的医学词汇数量
import pandas as pd
from openpyxl import load_workbook
import openpyxl

# ——————读字典——————
f = open("mydicts.txt",'r',encoding='utf-8')             # 返回一个文件对象
line = f.readline()             # 调用文件的 readline()方法
medworld=[]
while line:
# 后面跟 ',' 将忽略换行符
    a=line.strip('\n')

    medworld.append(a)
    line = f.readline()
print(medworld)
f.close()

# ——————读title——————
df = pd.read_excel("病例.xlsx",encoding='utf-8', usecols=[0],names=None)  # 读取项目名称列,不要列名
df_li = df.values.tolist()
result_keyw = []
for s_li in df_li:
    result_keyw.append(s_li[0])

# ——————配对——————
freq=[] #词频
wrds=[] #匹配词
ind=0   #序列
for i in result_keyw:
    i=str(i)
    wrd=''
    ii=0 #匹配中词语个数
    for j in medworld:
        if i.find(j)!=-1:
            ii+=1
            wrd += j + '  '
            print(j) #匹配的词语
    print(i,ii)
    # print(ind)
    # ind+=1
    wrds.append(wrd)
    freq.append(ii)
    print('-----------------------------------------------------------------')

# 写入xlsx
# for i in range(0,len(freq)):
#     wb = load_workbook("123.xlsx")#生成一个已存在的wookbook对象
#     wb1 = wb.active#激活sheet
#     wb1.cell(i+2,3,freq[i])#往sheet中的第二行第二列写入‘pass2’的数据
#     wb.save("123.xlsx")#保存
#     print('done',i)

# 写csv
dataframe = pd.DataFrame(
    {'result': freq,'words':wrds}
)
dataframe.to_csv("result.csv", index=False, encoding='utf_8_sig')
